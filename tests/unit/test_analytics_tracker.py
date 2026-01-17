"""
Unit tests for AnalyticsTracker

Tests analytics tracking sheet generation including:
- CSV generation
- Excel (XLSX) generation
- Column formatting
- Formula creation
- Summary sheet generation
"""

import csv
import pytest
from unittest.mock import patch
from datetime import date, time
from src.utils.analytics_tracker import (
    AnalyticsTracker,
    get_default_analytics_tracker,
    OPENPYXL_AVAILABLE,
)
from src.models.post import Post
from src.models.posting_schedule import PostingSchedule, ScheduledPost
from src.models.client_brief import Platform


class TestAnalyticsTrackerInit:
    """Test AnalyticsTracker initialization"""

    def test_init_creates_instance(self):
        """Test that AnalyticsTracker initializes"""
        tracker = AnalyticsTracker()
        assert tracker is not None

    def test_columns_defined(self):
        """Test that column definitions exist"""
        tracker = AnalyticsTracker()
        assert len(tracker.COLUMNS) > 0
        assert "Post #" in tracker.COLUMNS
        assert "Template" in tracker.COLUMNS
        assert "Impressions" in tracker.COLUMNS


class TestCreateTrackingSheetCsv:
    """Test CSV tracking sheet creation"""

    @pytest.fixture
    def tracker(self):
        """Create AnalyticsTracker instance"""
        return AnalyticsTracker()

    @pytest.fixture
    def sample_posts(self):
        """Sample posts for tracking"""
        return [
            Post(
                content="Test post 1 content",
                template_id=1,
                template_name="Problem Recognition",
                client_name="Test Client",
                word_count=20,
                has_cta=True,
                target_platform=Platform.LINKEDIN,
            ),
            Post(
                content="Test post 2 content",
                template_id=2,
                template_name="Statistic + Insight",
                client_name="Test Client",
                word_count=25,
                has_cta=False,
                target_platform=Platform.TWITTER,
            ),
        ]

    @pytest.fixture
    def sample_schedule(self):
        """Sample posting schedule"""
        from src.models.posting_schedule import DayOfWeek

        return PostingSchedule(
            client_name="Test Client",
            start_date=date(2025, 1, 1),
            end_date=date(2025, 1, 31),
            total_weeks=5,
            posts_per_week=6,
            scheduled_posts=[
                ScheduledPost(
                    post_id=1,
                    post_title="Test Post 1",
                    post_excerpt="Problem recognition post excerpt",
                    platform="LinkedIn",
                    scheduled_date=date(2025, 1, 1),
                    scheduled_time=time(9, 0),
                    day_of_week=DayOfWeek.WEDNESDAY,
                    week_number=1,
                ),
                ScheduledPost(
                    post_id=2,
                    post_title="Test Post 2",
                    post_excerpt="Statistic and insight post excerpt",
                    platform="Twitter",
                    scheduled_date=date(2025, 1, 2),
                    scheduled_time=time(10, 0),
                    day_of_week=DayOfWeek.THURSDAY,
                    week_number=1,
                ),
            ],
        )

    def test_create_csv_success(self, tracker, sample_posts, sample_schedule, tmp_path):
        """Test successful CSV creation"""
        output_path = tmp_path / "tracker.csv"

        result_path = tracker.create_tracking_sheet(
            posts=sample_posts,
            schedule=sample_schedule,
            output_path=output_path,
            format="csv",
        )

        assert result_path == output_path
        assert output_path.exists()

        # Verify CSV content
        with open(output_path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)

            # Check header
            assert rows[0] == tracker.COLUMNS

            # Check data rows
            assert len(rows) == 3  # Header + 2 posts
            assert rows[1][0] == "1"  # Post #
            assert rows[1][1] == "Problem Recognition"  # Template
            assert rows[2][0] == "2"

    def test_create_csv_creates_parent_dir(self, tracker, sample_posts, sample_schedule, tmp_path):
        """Test that parent directories are created"""
        output_path = tmp_path / "subdir" / "tracker.csv"

        result_path = tracker.create_tracking_sheet(
            posts=sample_posts,
            schedule=sample_schedule,
            output_path=output_path,
            format="csv",
        )

        assert output_path.parent.exists()
        assert result_path.exists()

    def test_create_csv_includes_schedule_dates(
        self, tracker, sample_posts, sample_schedule, tmp_path
    ):
        """Test that CSV includes scheduled dates"""
        output_path = tmp_path / "tracker.csv"

        tracker.create_tracking_sheet(
            posts=sample_posts,
            schedule=sample_schedule,
            output_path=output_path,
            format="csv",
        )

        with open(output_path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)

            # Check scheduled date in first post
            assert "2025-01-01" in rows[1][3]  # Scheduled Date column
            assert "09:00" in rows[1][4]  # Scheduled Time column

    def test_create_csv_handles_missing_schedule(self, tracker, sample_posts, tmp_path):
        """Test CSV creation with posts not in schedule"""
        output_path = tmp_path / "tracker.csv"

        # Empty schedule
        empty_schedule = PostingSchedule(
            client_name="Test Client",
            start_date=date(2025, 1, 1),
            end_date=date(2025, 1, 31),
            total_weeks=5,
            posts_per_week=6,
            scheduled_posts=[],
        )

        tracker.create_tracking_sheet(
            posts=sample_posts,
            schedule=empty_schedule,
            output_path=output_path,
            format="csv",
        )

        with open(output_path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)

            # Schedule columns should be empty
            assert rows[1][3] == ""  # No scheduled date
            assert rows[1][4] == ""  # No scheduled time

    def test_create_csv_includes_cta_status(self, tracker, sample_posts, sample_schedule, tmp_path):
        """Test that CSV includes CTA status"""
        output_path = tmp_path / "tracker.csv"

        tracker.create_tracking_sheet(
            posts=sample_posts,
            schedule=sample_schedule,
            output_path=output_path,
            format="csv",
        )

        with open(output_path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)

            # Check CTA column
            assert "Yes" in rows[1][6]  # Post 1 has CTA
            assert "No" in rows[2][6]  # Post 2 no CTA

    def test_create_csv_includes_keywords(self, tracker, sample_schedule, tmp_path):
        """Test that CSV includes keywords if present"""
        posts = [
            Post(
                content="Test content",
                template_id=1,
                template_name="Test",
                client_name="Test",
                word_count=10,
            ),
        ]
        # Add keywords_used attribute
        posts[0].__dict__["keywords_used"] = ["keyword1", "keyword2", "keyword3", "keyword4"]

        output_path = tmp_path / "tracker.csv"

        tracker.create_tracking_sheet(
            posts=posts,
            schedule=sample_schedule,
            output_path=output_path,
            format="csv",
        )

        with open(output_path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)

            # Check keywords column (should limit to first 3)
            assert "keyword1" in rows[1][7]
            assert "keyword2" in rows[1][7]
            assert "keyword3" in rows[1][7]


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
class TestCreateTrackingSheetXlsx:
    """Test Excel tracking sheet creation"""

    @pytest.fixture
    def tracker(self):
        """Create AnalyticsTracker instance"""
        return AnalyticsTracker()

    @pytest.fixture
    def sample_posts(self):
        """Sample posts for tracking"""
        return [
            Post(
                content="Test post content",
                template_id=1,
                template_name="Test Template",
                client_name="Test Client",
                word_count=15,
                has_cta=True,
            ),
        ]

    @pytest.fixture
    def sample_schedule(self):
        """Sample posting schedule"""
        from src.models.posting_schedule import DayOfWeek

        return PostingSchedule(
            client_name="Test Client",
            start_date=date(2025, 1, 1),
            end_date=date(2025, 1, 31),
            total_weeks=5,
            posts_per_week=6,
            scheduled_posts=[
                ScheduledPost(
                    post_id=1,
                    post_title="Test Post",
                    post_excerpt="Test excerpt",
                    platform="LinkedIn",
                    scheduled_date=date(2025, 1, 1),
                    scheduled_time=time(9, 0),
                    day_of_week=DayOfWeek.WEDNESDAY,
                    week_number=1,
                ),
            ],
        )

    def test_create_xlsx_success(self, tracker, sample_posts, sample_schedule, tmp_path):
        """Test successful Excel creation"""
        output_path = tmp_path / "tracker.xlsx"

        result_path = tracker.create_tracking_sheet(
            posts=sample_posts,
            schedule=sample_schedule,
            output_path=output_path,
            format="xlsx",
        )

        assert result_path == output_path
        assert output_path.exists()

    def test_create_xlsx_has_data_sheet(self, tracker, sample_posts, sample_schedule, tmp_path):
        """Test that Excel has data sheet"""
        from openpyxl import load_workbook

        output_path = tmp_path / "tracker.xlsx"

        tracker.create_tracking_sheet(
            posts=sample_posts,
            schedule=sample_schedule,
            output_path=output_path,
            format="xlsx",
        )

        wb = load_workbook(output_path)
        assert "Post Tracker" in wb.sheetnames

    def test_create_xlsx_has_summary_sheet(self, tracker, sample_posts, sample_schedule, tmp_path):
        """Test that Excel has summary sheet"""
        from openpyxl import load_workbook

        output_path = tmp_path / "tracker.xlsx"

        tracker.create_tracking_sheet(
            posts=sample_posts,
            schedule=sample_schedule,
            output_path=output_path,
            format="xlsx",
        )

        wb = load_workbook(output_path)
        assert "Summary" in wb.sheetnames

    def test_create_xlsx_header_formatting(self, tracker, sample_posts, sample_schedule, tmp_path):
        """Test that Excel header has formatting"""
        from openpyxl import load_workbook

        output_path = tmp_path / "tracker.xlsx"

        tracker.create_tracking_sheet(
            posts=sample_posts,
            schedule=sample_schedule,
            output_path=output_path,
            format="xlsx",
        )

        wb = load_workbook(output_path)
        ws = wb["Post Tracker"]

        # Check header formatting
        header_cell = ws["A1"]
        assert header_cell.font.bold is True
        assert header_cell.fill.start_color.rgb is not None

    def test_create_xlsx_includes_formulas(self, tracker, sample_posts, sample_schedule, tmp_path):
        """Test that Excel summary includes formulas"""
        from openpyxl import load_workbook

        output_path = tmp_path / "tracker.xlsx"

        tracker.create_tracking_sheet(
            posts=sample_posts,
            schedule=sample_schedule,
            output_path=output_path,
            format="xlsx",
        )

        wb = load_workbook(output_path)
        ws = wb["Summary"]

        # Check for formula cells
        total_posts_formula = ws["B3"].value
        assert "COUNTA" in total_posts_formula

        total_impressions_formula = ws["B5"].value
        assert "SUM" in total_impressions_formula


class TestCreateTrackingSheetFallback:
    """Test fallback behavior when openpyxl not available"""

    @pytest.fixture
    def tracker(self):
        """Create AnalyticsTracker instance"""
        return AnalyticsTracker()

    @pytest.fixture
    def sample_posts(self):
        """Sample posts"""
        return [
            Post(
                content="Test",
                template_id=1,
                template_name="Test",
                client_name="Test",
                word_count=10,
            ),
        ]

    @pytest.fixture
    def sample_schedule(self):
        """Sample schedule"""
        return PostingSchedule(
            client_name="Test",
            start_date=date(2025, 1, 1),
            end_date=date(2025, 1, 31),
            total_weeks=5,
            posts_per_week=6,
            scheduled_posts=[],
        )

    def test_xlsx_fallback_to_csv_when_openpyxl_unavailable(
        self, tracker, sample_posts, sample_schedule, tmp_path
    ):
        """Test that XLSX format falls back to CSV when openpyxl unavailable"""
        output_path = tmp_path / "tracker.xlsx"

        # Mock OPENPYXL_AVAILABLE to False
        with patch("src.utils.analytics_tracker.OPENPYXL_AVAILABLE", False):
            result_path = tracker.create_tracking_sheet(
                posts=sample_posts,
                schedule=sample_schedule,
                output_path=output_path,
                format="xlsx",
            )

            # Should create CSV instead
            assert result_path.suffix == ".csv"
            assert result_path.exists()


class TestCreateTrackingSheetValidation:
    """Test input validation"""

    @pytest.fixture
    def tracker(self):
        """Create AnalyticsTracker instance"""
        return AnalyticsTracker()

    @pytest.fixture
    def sample_posts(self):
        """Sample posts"""
        return [
            Post(
                content="Test",
                template_id=1,
                template_name="Test",
                client_name="Test",
                word_count=10,
            ),
        ]

    @pytest.fixture
    def sample_schedule(self):
        """Sample schedule"""
        return PostingSchedule(
            client_name="Test",
            start_date=date(2025, 1, 1),
            end_date=date(2025, 1, 31),
            total_weeks=5,
            posts_per_week=6,
            scheduled_posts=[],
        )

    def test_unsupported_format_raises_error(
        self, tracker, sample_posts, sample_schedule, tmp_path
    ):
        """Test that unsupported format raises ValueError"""
        output_path = tmp_path / "tracker.pdf"

        with pytest.raises(ValueError, match="Unsupported format"):
            tracker.create_tracking_sheet(
                posts=sample_posts,
                schedule=sample_schedule,
                output_path=output_path,
                format="pdf",
            )


class TestGetDefaultAnalyticsTracker:
    """Test get_default_analytics_tracker function"""

    def test_get_default_returns_instance(self):
        """Test that get_default_analytics_tracker returns instance"""
        tracker = get_default_analytics_tracker()
        assert isinstance(tracker, AnalyticsTracker)

    def test_get_default_returns_same_instance(self):
        """Test that get_default_analytics_tracker returns singleton"""
        tracker1 = get_default_analytics_tracker()
        tracker2 = get_default_analytics_tracker()
        assert tracker1 is tracker2


class TestIntegration:
    """Integration tests for analytics tracker"""

    def test_full_csv_generation_pipeline(self, tmp_path):
        """Test complete CSV generation pipeline"""
        tracker = AnalyticsTracker()

        posts = [
            Post(
                content="First post about collaboration and teamwork.",
                template_id=1,
                template_name="Problem Recognition",
                client_name="Test Client",
                word_count=6,
                has_cta=True,
                target_platform=Platform.LINKEDIN,
            ),
            Post(
                content="Second post with statistics and insights.",
                template_id=2,
                template_name="Statistic + Insight",
                client_name="Test Client",
                word_count=6,
                has_cta=False,
                target_platform=Platform.TWITTER,
            ),
        ]

        from src.models.posting_schedule import DayOfWeek

        schedule = PostingSchedule(
            client_name="Test Client",
            start_date=date(2025, 1, 1),
            end_date=date(2025, 1, 31),
            total_weeks=5,
            posts_per_week=6,
            scheduled_posts=[
                ScheduledPost(
                    post_id=1,
                    post_title="Test Post 1",
                    post_excerpt="Problem recognition excerpt",
                    platform="LinkedIn",
                    scheduled_date=date(2025, 1, 1),
                    scheduled_time=time(9, 0),
                    day_of_week=DayOfWeek.WEDNESDAY,
                    week_number=1,
                ),
                ScheduledPost(
                    post_id=2,
                    post_title="Test Post 2",
                    post_excerpt="Statistic and insight excerpt",
                    platform="Twitter",
                    scheduled_date=date(2025, 1, 2),
                    scheduled_time=time(14, 30),
                    day_of_week=DayOfWeek.THURSDAY,
                    week_number=1,
                ),
            ],
        )

        output_path = tmp_path / "full_tracker.csv"

        result = tracker.create_tracking_sheet(
            posts=posts,
            schedule=schedule,
            output_path=output_path,
            format="csv",
        )

        assert result.exists()

        # Verify complete CSV structure
        with open(result, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)

            assert len(rows) == 3  # Header + 2 posts
            assert len(rows[0]) == len(tracker.COLUMNS)

            # Verify post 1 data
            assert rows[1][0] == "1"
            assert rows[1][1] == "Problem Recognition"
            assert "2025-01-01" in rows[1][3]
            assert "09:00" in rows[1][4]
            assert "Yes" in rows[1][6]  # Has CTA

            # Verify post 2 data
            assert rows[2][0] == "2"
            assert rows[2][1] == "Statistic + Insight"
            assert "2025-01-02" in rows[2][3]
            assert "14:30" in rows[2][4]
            assert "No" in rows[2][6]  # No CTA

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_full_xlsx_generation_pipeline(self, tmp_path):
        """Test complete Excel generation pipeline"""
        from openpyxl import load_workbook

        tracker = AnalyticsTracker()

        posts = [
            Post(
                content="Test post content.",
                template_id=1,
                template_name="Test Template",
                client_name="Test Client",
                word_count=3,
                has_cta=True,
            ),
        ]

        from src.models.posting_schedule import DayOfWeek

        schedule = PostingSchedule(
            client_name="Test Client",
            start_date=date(2025, 1, 1),
            end_date=date(2025, 1, 31),
            total_weeks=5,
            posts_per_week=6,
            scheduled_posts=[
                ScheduledPost(
                    post_id=1,
                    post_title="Test Post",
                    post_excerpt="Test excerpt",
                    platform="LinkedIn",
                    scheduled_date=date(2025, 1, 1),
                    scheduled_time=time(9, 0),
                    day_of_week=DayOfWeek.WEDNESDAY,
                    week_number=1,
                ),
            ],
        )

        output_path = tmp_path / "full_tracker.xlsx"

        result = tracker.create_tracking_sheet(
            posts=posts,
            schedule=schedule,
            output_path=output_path,
            format="xlsx",
        )

        assert result.exists()

        # Verify workbook structure
        wb = load_workbook(result)
        assert "Post Tracker" in wb.sheetnames
        assert "Summary" in wb.sheetnames

        # Verify data sheet
        ws_data = wb["Post Tracker"]
        assert ws_data["A1"].value == "Post #"
        assert ws_data["A2"].value == 1
        assert ws_data["B2"].value == "Test Template"

        # Verify summary sheet has formulas
        ws_summary = wb["Summary"]
        assert "COUNTA" in ws_summary["B3"].value
        assert "SUM" in ws_summary["B5"].value
